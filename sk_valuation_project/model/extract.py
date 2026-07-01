"""Recalc FCFF_model_lines.xlsx via LibreOffice and print key Consolidated values."""
import os, sys, subprocess, glob
sys.path.insert(0, "/mnt/skills/public/xlsx/scripts")
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "output", "FCFF_model_lines.xlsx")
YEARS = list(range(2026, 2036))

def recalc():
    subprocess.run(["pkill", "-9", "soffice"], capture_output=True)
    import time; time.sleep(1)
    r = subprocess.run([sys.executable, "/mnt/skills/public/xlsx/scripts/recalc.py", XLSX, "180"],
                       capture_output=True, text=True)
    print(r.stdout[-400:]);
    if r.returncode != 0: print("RECALC ERR", r.stderr[-400:])

def findrow(ws, text):
    for row in ws.iter_rows(min_col=2, max_col=2):
        c = row[0]
        if c.value and text in str(c.value):
            return c.row
    return None

def rowvals(ws, rownum):
    return [ws.cell(rownum, 4+i).value for i in range(len(YEARS))]

def main():
    recalc()
    wb = load_workbook(XLSX, data_only=True)
    cs = wb["Consolidated"]
    def show(label, key_text, single=False):
        rn = findrow(cs, key_text)
        if rn is None: print(f"  {label}: (not found '{key_text}')"); return
        if single:
            print(f"  {label}: {cs.cell(rn,4).value}")
        else:
            vals = rowvals(cs, rn)
            print(f"  {label}: " + " ".join(f"{v:.0f}" if isinstance(v,(int,float)) else str(v) for v in vals))
    print("YEARS:            " + " ".join(str(y) for y in YEARS))
    show("EBITDA ex-AMPC", "EBITDA ex-AMPC (멀티플용)")
    show("EBITDA ex-AMPC %","EBITDA ex-AMPC margin")
    show("EBITDA inclAMPC", "EBITDA incl. AMPC")
    show("Total revenue",   "Total revenue")
    show("Consol FCFF",     "Consolidated FCFF")
    show("SK FCFE",         "보통주(SK) FCFE 합계")
    print("---- summary ----")
    show("Fwd exAMPC 2029E","Forward EBITDA ex-AMPC (2029E)", single=True)
    show("Run-rate exAMPC", "Run-rate EBITDA ex-AMPC", single=True)
    show("Enterprise Value","Enterprise Value (EV)", single=True)
    show("Equity (DCF)",    "Equity Value (DCF)", single=True)
    show("Equity (FCFE)",   "Equity Value (FCFE", single=True)
    # error scan
    errs=0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value,str) and c.value.startswith("#"): errs+=1
    print(f"---- total_errors={errs} ----")

if __name__ == "__main__":
    main()
