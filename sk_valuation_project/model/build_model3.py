from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1B1714"; RED="C1121F"; SAND="EFE9DC"; LINEC="CFC6B4"
BLUE="0000FF"; BLACK="000000"; GREEN="008000"
P1C="1F3A63"; P2C="7A2230"; POUCH="2E6FA8"; PRIS="A6552E"
YEL="FFF6CC"; BAtint="EAF0F7"; OTtint="F6ECEE"
F="Arial"
def fnt(sz=10,b=False,color=BLACK,it=False): return Font(name=F,size=sz,bold=b,color=color,italic=it)
fnav=PatternFill("solid",fgColor=NAVY); fsand=PatternFill("solid",fgColor=SAND); fred=PatternFill("solid",fgColor=RED)
fp1=PatternFill("solid",fgColor=P1C); fp2=PatternFill("solid",fgColor=P2C); fyel=PatternFill("solid",fgColor=YEL)
fpouch=PatternFill("solid",fgColor=POUCH); fpris=PatternFill("solid",fgColor=PRIS)
thin=Side(style="thin",color=LINEC)
b_tb=Border(top=Side(style="thin",color=NAVY),bottom=Side(style="thin",color=NAVY))
b_t=Border(top=Side(style="thin",color=NAVY))
right=Alignment(horizontal="right"); center=Alignment(horizontal="center")
USD='$#,##0;($#,##0);"-"'; USD2='$#,##0.00;($#,##0.00);"-"'; NUM1='#,##0.0;(#,##0.0);"-"'; PCT='0.0%;(0.0%);"-"'
YEARS=list(range(2026,2036)); C0=4
def col(i): return get_column_letter(C0+i)

wb=Workbook()
A=wb.active; A.title="Assumptions"; A.sheet_view.showGridLines=False
A.column_dimensions['A'].width=2; A.column_dimensions['B'].width=34; A.column_dimensions['C'].width=13
for i in range(len(YEARS)): A.column_dimensions[col(i)].width=9.5

def sect(ws,r,t,fill=fnav):
    ws.cell(r,2,t).font=fnt(10,True,"FFFFFF")
    for cc in range(2,C0+len(YEARS)): ws.cell(r,cc).fill=fill
def yhead(ws,r,lab="per-line"):
    c=ws.cell(r,2,lab); c.font=fnt(9,True,NAVY); c.border=b_tb; ws.cell(r,3).border=b_tb
    for i,y in enumerate(YEARS):
        cc=ws.cell(r,C0+i,str(y)); cc.font=fnt(10,True,NAVY); cc.alignment=center; cc.border=b_tb
def lbl(ws,r,t,note="",bold=False,color=BLACK):
    ws.cell(r,2,t).font=fnt(10,bold,color)
    if note: ws.cell(r,3,note).font=fnt(8,False,"8C8473",it=True)
def blue(ws,r,i,v,fmt=NUM1,fill=None):
    c=ws.cell(r,C0+i,v); c.font=fnt(10,False,BLUE); c.number_format=fmt; c.alignment=right
    if fill: c.fill=fill
    return c
def form(ws,r,i,f,fmt=NUM1,bold=False,color=BLACK):
    c=ws.cell(r,C0+i,f); c.font=fnt(10,bold,color); c.number_format=fmt; c.alignment=right; return c

R={}
A.cell(1,2,"US LFP/ESS CELL PLATFORM — LINE-LEVEL FINANCIAL PROJECTION").font=fnt(15,True,NAVY)
A.cell(2,2,"Project [ ● ]  ·  Revenue → FCFF  ·  라인 단위  ·  Strictly Private & Confidential").font=fnt(9,False,"8C8473")
A.cell(3,2,"Blue=input  Black=formula  ★=가정값   |   ASP = 중국산 셀 + 미국 관세(58.4%) 하한 기반 도출 · SKBA 4라인 · SKOT 3라인").font=fnt(8.5,False,"8C8473",it=True)

r=5
# GLOBAL TOGGLE
A.cell(r,2,"GLOBAL TOGGLE").font=fnt(10,True,"FFFFFF")
for cc in range(2,C0+len(YEARS)): A.cell(r,cc).fill=fred
r+=1
A.cell(r,2,"Effective production factor (명판 대비) ★").font=fnt(10,True)
tg=A.cell(r,3,1.00); tg.font=fnt(11,True,BLUE); tg.number_format=PCT; tg.alignment=center; tg.fill=fyel
A.cell(r,4,"← 토글: 명판×이 값 = 실제 생산 (1.0 → util곡선=실효 가동률)").font=fnt(8,False,"8C8473",it=True)
R['eff']=r; r+=1
A.cell(r,2,"ASP 방식 토글 ★  (1=관세하한 셀 / 2=DC블록 $155)").font=fnt(10,True)
tg2=A.cell(r,3,2); tg2.font=fnt(11,True,BLUE); tg2.number_format='0'; tg2.alignment=center; tg2.fill=fyel
A.cell(r,4,"← 1: 셀 ASP(관세하한)  /  2: DC블록 시스템 ASP $155").font=fnt(8,False,"8C8473",it=True)
R['asp_mode']=r; r+=1
# AMPC schedule (법정, 공유)
A.cell(r,2,"AMPC 45X ($/kWh) — 셀+모듈(DC블록) · 법정 phase-out").font=fnt(10,True)
amp={2026:0,2027:45,2028:45,2029:45,2030:33.75,2031:22.5,2032:11.25,2033:0,2034:0,2035:0}
R['ampc']=r
for i,y in enumerate(YEARS): blue(A,r,i,amp[y],USD2)
A.cell(r,3,"'30=75%·'31=50%…").font=fnt(8,False,"8C8473",it=True)
r+=2

# ---------- ASP FLOOR via China + US tariffs ----------
sect(A,r,"ASP FLOOR — 중국산 LFP 셀 + 미국 관세 = 미국시장 ASP 하한",fp2); r+=1
A.cell(r,2,"논리: 중국산 셀가×(1+관세 스택)이 미국 ASP의 하한 — 미국산은 그 아래로 팔 이유 없음").font=fnt(8.5,True,P2C,it=True); r+=1
yhead(A,r,"$/kWh · %"); r+=1
# China cell price (input) — 2024 ~$73, declining base trajectory
lbl(A,r,"중국산 LFP 셀가 ($/kWh) ★","'25 내수~$40·수출~$50대"); R['cn_cell']=r
cn={2026:52,2027:50,2028:48,2029:46,2030:45,2031:44,2032:43,2033:43,2034:42,2035:42}
for i,y in enumerate(YEARS): blue(A,r,i,cn[y],USD)
r+=1
# tariff components (input, constant by default; user can ramp)
lbl(A,r,"  기본관세 MFN (%)","전세계 리튬이온 공통"); R['t_mfn']=r
for i in range(len(YEARS)): blue(A,r,i,0.034,PCT)
r+=1
lbl(A,r,"  상호관세 Reciprocal (%)","Sec.122 등"); R['t_recip']=r
for i in range(len(YEARS)): blue(A,r,i,0.10,PCT)
r+=1
lbl(A,r,"  펜타닐 IEEPA (%)",""); R['t_fent']=r
for i in range(len(YEARS)): blue(A,r,i,0.20,PCT)
r+=1
lbl(A,r,"  Section 301 (%)","'26~ 25% (비EV 배터리)"); R['t_301']=r
s301={2026:0.25,2027:0.25,2028:0.25,2029:0.25,2030:0.25,2031:0.25,2032:0.25,2033:0.25,2034:0.25,2035:0.25}
for i,y in enumerate(YEARS): blue(A,r,i,s301[y],PCT)
r+=1
lbl(A,r,"관세 합계 (%)","합산",bold=True); R['t_total']=r
for i in range(len(YEARS)):
    form(A,r,i,f"={col(i)}{R['t_mfn']}+{col(i)}{R['t_recip']}+{col(i)}{R['t_fent']}+{col(i)}{R['t_301']}",PCT,bold=True)
r+=1
lbl(A,r,"관세적용 중국산 = ASP 하한 ($/kWh)","중국셀가×(1+관세)",bold=True,color=P2C); R['asp_floor']=r
for i in range(len(YEARS)):
    form(A,r,i,f"={col(i)}{R['cn_cell']}*(1+{col(i)}{R['t_total']})",USD,bold=True,color=P2C)
r+=1
# US-made premium over floor (input %): how much above floor US producer prices
lbl(A,r,"미국산 프리미엄 (하한 대비 %) ★","FEOC-clean·안정공급 프리미엄"); R['us_prem']=r
for i in range(len(YEARS)): blue(A,r,i,0.20,PCT)
A.cell(r,3,"통념상 +20% (FEOC·국산 프리미엄)").font=fnt(8,False,"8C8473",it=True)
r+=1
lbl(A,r,"[방식1] 셀 ASP = 하한×(1+프리미엄)","토글=1일 때",color=POUCH); R['asp_cell']=r
for i in range(len(YEARS)):
    form(A,r,i,f"={col(i)}{R['asp_floor']}*(1+{col(i)}{R['us_prem']})",USD,color=POUCH)
r+=1
lbl(A,r,"[방식2] DC블록 ASP ($/kWh) ★","SK가 DC블록 시스템 공급"); R['asp_sys']=r
syp={2026:0,2027:155,2028:155,2029:153,2030:151,2031:149,2032:147,2033:145,2034:143,2035:141}
for i,y in enumerate(YEARS): blue(A,r,i,syp[y],USD)
A.cell(r,3,"$155 — 셀+모듈+컨테이너 DC블록 통합").font=fnt(8,False,"8C8473",it=True)
r+=1
lbl(A,r,"[방식2] 시스템 마진 프리미엄 (%p) ★","GM은 아래서 직접 설정→0"); R['sys_gm_prem']=r
for i in range(len(YEARS)): blue(A,r,i,0.0,PCT)
A.cell(r,3,"DC블록 GM은 동별 gross margin에 직접 반영").font=fnt(8,False,"8C8473",it=True)
r+=1
lbl(A,r,"적용 마진 프리미엄 (토글 반영)","",color=POUCH); R['gm_prem_eff']=r
for i in range(len(YEARS)):
    form(A,r,i,f"=IF($C${R['asp_mode']}=2,{col(i)}{R['sys_gm_prem']},0)",PCT,color=POUCH)
r+=1
lbl(A,r,"도출 ASP — 파우치 ($/kWh)","토글로 선택",bold=True,color=POUCH); R['asp_derived']=r
for i in range(len(YEARS)):
    form(A,r,i,f"=IF($C${R['asp_mode']}=2,{col(i)}{R['asp_sys']},{col(i)}{R['asp_cell']})",USD,bold=True,color=POUCH)
r+=2

# ---------- LINE DEFINITION TABLE ----------
# each line: capacity (GWh), COD year, utilization curve, formfactor
def line_rows(start_r, plant_tag, lines, tint, placeholder=False):
    """lines = list of dicts: name, cap, cod, util(dict by year or None), form"""
    tintfill=PatternFill("solid",fgColor=tint)
    rr=start_r
    out={}
    for ln in lines:
        lbl(A,rr,f"  {ln['name']}",f"{ln['form']} · COD {ln['cod'] if ln['cod'] else '★'}")
        # capacity row (single input cell C col, applied from COD)
        out[ln['key']]={}
        out[ln['key']]['cap_row']=rr
        for i,y in enumerate(YEARS):
            online = (ln['cod'] is not None and y>=ln['cod'])
            v = (ln['cap'] if online else 0)
            c=blue(A,rr,i,v,NUM1,fill=(fyel if placeholder else (tintfill if online else None)))
        rr+=1
        # utilization row
        lbl(A,rr,f"    └ 가동률 ★","")
        out[ln['key']]['util_row']=rr
        for i,y in enumerate(YEARS):
            uv = (ln['util'].get(y,0) if ln['util'] else 0)
            blue(A,rr,i,uv,PCT,fill=(fyel if placeholder else None))
        rr+=1
        # shipped row (formula)
        lbl(A,rr,f"    └ Shipped (GWh)","명판×가동×토글")
        out[ln['key']]['gwh_row']=rr
        for i in range(len(YEARS)):
            form(A,rr,i,f"={col(i)}{out[ln['key']]['cap_row']}*{col(i)}{out[ln['key']]['util_row']}*$C${R['eff']}",NUM1,bold=True)
        rr+=1
    return rr,out

# default util curves
# eff=1.0 이므로 아래 곡선이 곧 실효 가동률(명판 대비 출하). DC센터 수요 확보 → run-rate 85% 안착
ba_util_27={2027:.55,2028:.72,2029:.82,2030:.85,2031:.85,2032:.85,2033:.85,2034:.85,2035:.85}
ba_util_28={2028:.55,2029:.72,2030:.83,2031:.85,2032:.85,2033:.85,2034:.85,2035:.85}

sect(A,r,"1 · SKBA (공장1) — LFP 파우치 · 1-1동(5.3×2, 자체전극) + 1-2동(전극 외부구매)",fp1); r+=1
yhead(A,r); r+=1
A.cell(r,2,"  ▸ 공장1-1동 (파우치, '27, 라인당 5.3GW · 자체 전극)").font=fnt(9,True,P1C); r+=1
r,BA=line_rows(r,"ba",[
    {'name':"L1-1A",'key':'l11a','cap':5.3,'cod':2027,'util':ba_util_27,'form':"파우치"},
    {'name':"L1-1B",'key':'l11b','cap':5.3,'cod':2027,'util':ba_util_27,'form':"파우치"},
], BAtint)
A.cell(r,2,"  ▸ 공장1-2동 (파우치, '28, 전극 외부구매)").font=fnt(9,True,P1C); r+=1
r,BA2=line_rows(r,"ba",[
    {'name':"L1-2C",'key':'l12c','cap':5.3,'cod':2028,'util':ba_util_28,'form':"파우치"},
    {'name':"L1-2D",'key':'l12d','cap':5.3,'cod':2028,'util':ba_util_28,'form':"파우치"},
], BAtint)
BA.update(BA2)
# shipped by 동 (전극 sourcing 다름)
lbl(A,r,"1-1동 shipped (GWh)","자체 전극",bold=True); R['ba11_gwh']=r
for i in range(len(YEARS)):
    refs="+".join(f"{col(i)}{BA[k]['gwh_row']}" for k in ['l11a','l11b'])
    form(A,r,i,f"={refs}",NUM1,bold=True)
r+=1
lbl(A,r,"1-2동 shipped (GWh)","전극 외부구매",bold=True); R['ba12_gwh']=r
for i in range(len(YEARS)):
    refs="+".join(f"{col(i)}{BA[k]['gwh_row']}" for k in ['l12c','l12d'])
    form(A,r,i,f"={refs}",NUM1,bold=True)
r+=1
lbl(A,r,"SKBA shipped 합계 (GWh)","4개 라인 합",bold=True); R['ba_gwh']=r
for i in range(len(YEARS)):
    form(A,r,i,f"={col(i)}{R['ba11_gwh']}+{col(i)}{R['ba12_gwh']}",NUM1,bold=True)
r+=1
lbl(A,r,"ASP — 파우치 ($/kWh)","관세하한 도출 ASP 링크"); R['ba_asp']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['asp_derived']}",USD,color=POUCH)
r+=1
lbl(A,r,"  1-1동 gross margin (%) ★","자체 전극·ex-AMPC EBITDA 9% (지분 수익률 기준)"); R['ba11_gm']=r
for i,y in enumerate(YEARS): blue(A,r,i,(0 if y<2027 else 0.16),PCT)
r+=1
lbl(A,r,"  1-2동 gross margin (%) ★","전극 외부구매·통합마진 반영(ex-AMPC EBITDA 9%)"); R['ba12_gm']=r
for i,y in enumerate(YEARS): blue(A,r,i,(0 if y<2028 else 0.14),PCT)
r+=1
lbl(A,r,"1-1동 COGS ($/kWh)","ASP×(1−1-1마진)"); R['ba11_cogs']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['ba_asp']}*(1-{col(i)}{R['ba11_gm']}-{col(i)}{R['gm_prem_eff']})",USD)
r+=1
lbl(A,r,"1-2동 COGS ($/kWh)","전극 외부구매로 원가↑"); R['ba12_cogs']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['ba_asp']}*(1-{col(i)}{R['ba12_gm']}-{col(i)}{R['gm_prem_eff']})",USD)
r+=1
lbl(A,r,"Retrofit capex — 1-1동 ($mm) ★","전극공정 포함"); R['ba11_capex']=r
ba11cap={2026:200,2027:120,2028:0,2029:0,2030:0,2031:0,2032:0,2033:0,2034:0,2035:0}
for i,y in enumerate(YEARS): blue(A,r,i,ba11cap[y],USD)
r+=1
lbl(A,r,"Retrofit capex — 1-2동 ($mm) ★","전극 외부구매→경량 capex"); R['ba12_capex']=r
ba12cap={2026:0,2027:60,2028:40,2029:0,2030:0,2031:0,2032:0,2033:0,2034:0,2035:0}
for i,y in enumerate(YEARS): blue(A,r,i,ba12cap[y],USD)
r+=1
lbl(A,r,"SKBA capex 합계 ($mm)","",bold=True); R['ba_capex']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['ba11_capex']}+{col(i)}{R['ba12_capex']}",USD,bold=True)
r+=2

# ---------- SKOT (placeholders) ----------
sect(A,r,"2 · SKOT (공장2) — 각형 13.5×2 + 파우치 2.7 = 29.7 GWh · '28 가동",fp2); r+=1
A.cell(r,2,"라인당 캐파 입력값 (각형 13.5GW×2, 파우치 2.7GW) · COD 2028 가정").font=fnt(8.5,True,P2C,it=True); r+=1
yhead(A,r); r+=1
A.cell(r,2,"  ▸ 각형 라인 (prismatic) · 라인당 13.5GW").font=fnt(9,True,P2C); r+=1
ot_util_28={2028:.35,2029:.58,2030:.75,2031:.85,2032:.85,2033:.85,2034:.85,2035:.85}
r,OT=line_rows(r,"ot",[
    {'name':"각형 L1",'key':'opr1','cap':13.5,'cod':2028,'util':ot_util_28,'form':"각형"},
    {'name':"각형 L2",'key':'opr2','cap':13.5,'cod':2028,'util':ot_util_28,'form':"각형"},
], OTtint)
A.cell(r,2,"  ▸ 파우치 라인 (pouch) · 2.7GW").font=fnt(9,True,P2C); r+=1
r,OTp=line_rows(r,"ot",[
    {'name':"파우치 L3",'key':'opo3','cap':2.7,'cod':2028,'util':ot_util_28,'form':"파우치"},
], OTtint)
OT.update(OTp)
# SKOT split shipped by formfactor
lbl(A,r,"SKOT 각형 shipped (GWh)","각형 2라인",bold=True); R['ot_pr_gwh']=r
for i in range(len(YEARS)):
    refs="+".join(f"{col(i)}{OT[k]['gwh_row']}" for k in ['opr1','opr2'])
    form(A,r,i,f"={refs}",NUM1,bold=True)
r+=1
lbl(A,r,"SKOT 파우치 shipped (GWh)","파우치 1라인",bold=True); R['ot_po_gwh']=r
form_row=r
for i in range(len(YEARS)):
    form(A,r,i,f"={col(i)}{OT['opo3']['gwh_row']}",NUM1,bold=True)
r+=1
lbl(A,r,"  각형 할인 (파우치 대비 %) ★","각형 대용량·저가"); R['ot_pr_disc']=r
for i,y in enumerate(YEARS): blue(A,r,i,(0 if y<2028 else 0.04),PCT)
r+=1
lbl(A,r,"ASP — 각형 ($/kWh)","도출ASP×(1−할인)"); R['ot_pr_asp']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['asp_derived']}*(1-{col(i)}{R['ot_pr_disc']})",USD,color=PRIS)
r+=1
lbl(A,r,"  목표 각형 gross margin (%) ★",""); R['ot_pr_gm']=r
for i,y in enumerate(YEARS): blue(A,r,i,(0 if y<2028 else 0.15),PCT)
r+=1
lbl(A,r,"COGS — 각형 ($/kWh)","ASP×(1−마진)"); R['ot_pr_cogs']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['ot_pr_asp']}*(1-{col(i)}{R['ot_pr_gm']}-{col(i)}{R['gm_prem_eff']})",USD)
r+=1
lbl(A,r,"ASP — 파우치 ($/kWh)","도출ASP 링크(SKBA 동일)"); R['ot_po_asp']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['asp_derived']}",USD,color=POUCH)
r+=1
lbl(A,r,"  목표 파우치 gross margin (%) ★",""); R['ot_po_gm']=r
for i,y in enumerate(YEARS): blue(A,r,i,(0 if y<2028 else 0.15),PCT)
r+=1
lbl(A,r,"COGS — 파우치 ($/kWh)","ASP×(1−마진)"); R['ot_po_cogs']=r
for i in range(len(YEARS)): form(A,r,i,f"={col(i)}{R['ot_po_asp']}*(1-{col(i)}{R['ot_po_gm']}-{col(i)}{R['gm_prem_eff']})",USD)
r+=1
lbl(A,r,"Retrofit capex ($mm) ★","개조 ~$1.0bn"); R['ot_capex']=r
otcap={2026:200,2027:500,2028:300,2029:0,2030:0,2031:0,2032:0,2033:0,2034:0,2035:0}
for i,y in enumerate(YEARS): blue(A,r,i,otcap[y],USD)
r+=2

# ---------- SHARED ----------
sect(A,r,"3 · SHARED — OPEX / CAPEX / WC / DCF"); r+=1
def kv(r,label,val,fmt=PCT,note=""):
    A.cell(r,2,label).font=fnt(10)
    c=A.cell(r,3,val); c.font=fnt(10,False,BLUE); c.number_format=fmt; c.alignment=right
    if note: A.cell(r,4,note).font=fnt(8,False,"8C8473",it=True)
    return r
R['sga']=kv(r,"SG&A (% of revenue) ★",0.06); r+=1
R['dep']=kv(r,"D&A (% of gross PP&E) ★",0.0667,PCT,"≈15y"); r+=1
R['mcapex']=kv(r,"Maintenance capex (% rev) ★",0.03); r+=1
R['capex_grant']=kv(r,"CapEx 정부 그랜트 (% of growth capex) ★",0.10,PCT,"설비보조금→SK 출자 경감"); r+=1
R['ba_base_ebitda']=kv(r,"SKBA 기저 EBITDA floor (pre-SOP, $mm/yr) ★",200,USD,"양산 전(2026~) 기존 사업 하한 · 개조분이 이보다 크면 미적용(run-rate 유지)"); r+=1
R['nwc']=kv(r,"Net working capital (% rev) ★",0.15); r+=1
R['tax']=kv(r,"Cash tax rate ★",0.23,PCT,"AMPC 비과세"); r+=1
R['nol_ba']=kv(r,"SKBA 기초 NOL (누적결손) ★",4000,USD,"SK온 출자분 중 누적손실(추정·DART 확인要)"); r+=1
R['nol_ot']=kv(r,"SKOT 기초 NOL ★",0,USD,"분할 신설 → 0"); r+=1
R['nol_lim']=kv(r,"NOL 공제 한도 (과세소득 %) ★",0.80,PCT,"TCJA: 80%까지만 상계"); r+=1
R['wacc']=kv(r,"WACC ★",0.105); r+=1
R['coe']=kv(r,"Cost of equity (FCFE 할인) ★",0.13,PCT,"레버리지 반영 Ke"); r+=1
R['tgr']=kv(r,"Terminal growth (TGR) ★",0.02); r+=1
R['netdebt']=kv(r,"Net debt ($mm) ★",6400,USD,"BA/OT 합산 순차입금 = BA 3,700 + OT 2,600"); r+=1
R['midyr']=kv(r,"Mid-year convention",0.5,'0.0'); r+=2
# ---- 자본구조 (FCFE용) — 하이브리드: HoldCo 우산 + 법인별 격벽 ----
sect(A,r,"자본구조 — FCFE용 (하이브리드: HoldCo 우산 + 법인별 ring-fence)",fp2); r+=1
A.cell(r,2,"위(HoldCo)=통합 조달 / 아래(SPV)=격리·non-recourse · waiver는 OT 후순위 출자에만 좁게 사용").font=fnt(8.5,True,P2C,it=True); r+=1
# --- HoldCo 레벨 (공통 조달) ---
A.cell(r,2,"  ▸ HoldCo 레벨 (정부앵커·전략·인프라 크레딧 통합 조달)").font=fnt(9,True,P2C); r+=1
R['hc_mezz']=kv(r,"HoldCo · 인프라 크레딧/메자닌 ($mm) ★",0,USD,"그랜트우선 구조: HoldCo 크레딧 레이어 제거"); r+=1
R['hc_rate']=kv(r,"HoldCo · 크레딧 비용률 ★",0.09,PCT,"통합조달 → 단독보다 저렴"); r+=1
# --- SKBA SPV (DOE-clean) ---
A.cell(r,2,"  ▸ SKBA SPV (DOE-clean · 외부지분 ≤19.9%)").font=fnt(9,True,P1C); r+=1
R['ba_debt']=kv(r,"신규 비소구부채 ($mm) ★",852,USD,"그랜트우선: growth capex $1.42bn × 60% (BA 일괄조달)"); r+=1
R['ba_rate']=kv(r,"비소구부채 이자율 ★",0.065,PCT,"비소구 → 소폭 가산"); r+=1
R['ba_hc']=kv(r,"SKBA · HoldCo 후순위 출자 ($mm) ★",0,USD,"그랜트우선 구조: 제거"); r+=1
# --- SKOT SPV (ring-fenced) ---
A.cell(r,2,"  ▸ SKOT SPV (DOE ring-fenced · waiver 전제)").font=fnt(9,True,P2C); r+=1
R['doe_bal']=kv(r,"SKOT · DOE loan 잔액 ($mm) ★",4000,USD,"이자만·원금상환 없음"); r+=1
R['doe_rate']=kv(r,"SKOT · DOE 이자율 ★",0.05,PCT,"이자만 5.0%"); r+=1
R['pref_bal']=kv(r,"SKOT · 메자닌 ($mm) — 제거 ★",0,USD,"메자닌 제거: OT는 BA 증자(equity-down)로 자금조달"); r+=1
R['pref_rate']=kv(r,"SKOT · 메자닌 배당률 (inert) ★",0.07,PCT,"메자닌 제거로 미사용 (bal=0)"); r+=1
R['pref_amort_start']=kv(r,"SKOT · 메자닌 상환개시 (연도) ★",2029,'0',"(A)현금발생後 거치상환"); r+=1
R['pref_mat']=kv(r,"SKOT · 메자닌 만기 (연도) ★",2032,'0',"(A)AMPC 창내 상환"); r+=1
R['pref_kick']=kv(r,"SKOT · 에쿼티 키커 워런트 (%) ★",0.05,PCT,"(B)OT SPV 지분, 현금정산형"); r+=1
R['ot_hc']=kv(r,"SKOT · HoldCo 후순위 출자 ($mm) ★",0,USD,"그랜트우선 구조: 제거 (OT는 BA 증자)"); r+=2
A.cell(r,2,"FCFE = FCFF − 이자×(1−세금) − 우선배당(세후) − HoldCo 크레딧비용 배분 ± 순차입. DOE 이자만(원금상환 0).").font=fnt(8,False,"8C8473",it=True); r+=1
A.cell(r,2,"HoldCo 크레딧비용은 후순위 출자액 비례로 BA·OT에 배분(세후, tax-shield 無 가정 — 메자닌 배당성격).").font=fnt(8,False,"8C8473",it=True); r+=2
R['skba_nav']=kv(r,"SKBA 순자산 NAV ($mm) ★",2000,USD,"추정: 출자~4.8조−누적결손, DART 확인要"); 
A.cell(r,3).fill=fyel; r+=1
R['skot_nav']=kv(r,"SKOT 순자산 NAV ($mm) ★",3000,USD,"분할 PPA ~$3bn (DOE $4bn 차감後 순자산)"); r+=2
A.cell(r,2,"DART: SKBA 순자산·손익은 SK이노베이션 연결 사업보고서 종속기업 명세 공시(예 '22 -4,608억원). SKOT는 분할 신설→입력. AMPC 옵션A(셀 $35·법정 phase-out, 모듈 미반영).").font=fnt(8,False,"8C8473",it=True)
r+=1
A.cell(r,2,"NAV: SKBA 순자산 ~$2.0bn (추정: SK온 출자 ~4.8조원 − 누적결손; DART 자본총계 확인 필요) · SKOT $3.0bn (분할 PPA). SOTP 교차검증용.").font=fnt(8,False,RED,it=True)

AR=lambda k,i=None: (f"Assumptions!{col(i)}{R[k]}" if i is not None else f"Assumptions!$C${R[k]}")

# ============================================================
# PLANT P&L sheets (SKBA single formfactor; SKOT dual formfactor)
# ============================================================
def H(ws,r):
    ws.cell(r,2,"($mm)").font=fnt(9,True,NAVY); ws.cell(r,2).border=b_tb; ws.cell(r,3).border=b_tb
    for i,y in enumerate(YEARS):
        c=ws.cell(r,C0+i,str(y)); c.font=fnt(10,True,NAVY); c.alignment=center; c.border=b_tb
def Srow(ws,r,t,fill):
    ws.cell(r,2,t).font=fnt(10,True,"FFFFFF")
    for cc in range(2,C0+len(YEARS)): ws.cell(r,cc).fill=PatternFill("solid",fgColor=fill)

def plant_common(ws,PROW,rev_row,ampc_gwh_keys,capex_key,head,nol_key,base_key=None):
    """build cost->EBIT->FCFF given a revenue row already placed; ampc on total shipped"""
    r=PROW['after_rev']
    Srow(ws,r,"COST & EBIT",head); r+=1; H(ws,r); r+=1
    # total cash COGS placed by caller as 'cogs'
    def L(r,label,key,fnc,bold=False,fmt=USD,top=False,fill=None,pct=False):
        lc=ws.cell(r,2,label); lc.font=fnt(10,bold,NAVY if bold else BLACK); PROW[key]=r
        for i in range(len(YEARS)):
            c=ws.cell(r,C0+i,fnc(i)); c.font=fnt(10,bold,BLACK); c.number_format=(PCT if pct else fmt); c.alignment=right
            if fill: c.fill=fill
            if top: c.border=b_t
        if fill: lc.fill=fill
        if top: lc.border=b_t
        return r
    L(r,"Gross profit (pre-AMPC)","gp",lambda i:f"={col(i)}{PROW['rev']}+{col(i)}{PROW['cogs']}",bold=True,top=True); r+=1
    L(r,"SG&A","sga",lambda i:f"=-{col(i)}{PROW['rev']}*{AR('sga')}"); r+=1
    PROW['ppe']=r; ws.cell(r,2,"Gross PP&E (cum.)").font=fnt(10)
    for i in range(len(YEARS)):
        f=(f"={AR(capex_key,0)}" if i==0 else f"={col(i-1)}{PROW['ppe']}+{AR(capex_key,i)}")
        c=ws.cell(r,C0+i,f); c.font=fnt(10); c.number_format=USD; c.alignment=right
    r+=1
    L(r,"D&A","da",lambda i:f"=-{col(i)}{PROW['ppe']}*{AR('dep')}"); r+=1
    if base_key:
        # pre-SOP floor: ex-AMPC EBITDA(=gp+sga)가 기저치 미만이면 그 차액만 보정 → run-rate(개조분>기저)엔 영향 0
        L(r,"기저 EBITDA floor 보정 (pre-SOP)","base",
          lambda i:f"=MAX(0,{AR(base_key)}-({col(i)}{PROW['gp']}+{col(i)}{PROW['sga']}))"); r+=1
    base_term=(lambda i:f"+{col(i)}{PROW['base']}") if base_key else (lambda i:"")
    L(r,"EBIT (pre-AMPC)","ebitpre",lambda i:f"={col(i)}{PROW['gp']}+{col(i)}{PROW['sga']}+{col(i)}{PROW['da']}{base_term(i)}",bold=True,top=True); r+=1
    L(r,"  margin pre-AMPC","ebitprem",lambda i:f"=IF({col(i)}{PROW['rev']}=0,0,{col(i)}{PROW['ebitpre']}/{col(i)}{PROW['rev']})",pct=True); r+=1
    # AMPC on total shipped (sum of given gwh refs)
    L(r,"Plus: AMPC 45X","ampc",lambda i:"=("+"+".join(ampc_gwh_keys(i))+f")*{AR('ampc',i)}"); r+=1
    L(r,"EBIT (incl. AMPC)","ebit",lambda i:f"={col(i)}{PROW['ebitpre']}+{col(i)}{PROW['ampc']}",bold=True,top=True); r+=2
    Srow(ws,r,"UNLEVERED FCFF",head); r+=1; H(ws,r); r+=1
    # --- NOL (누적결손 이월) · noleop는 nolbop +4행 뒤 ---
    L(r,"과세소득 pre-NOL","txpre",lambda i:f"=MAX(0,{col(i)}{PROW['ebitpre']})"); r+=1
    nolbop_r=r
    L(r,"NOL 기초잔액","nolbop",lambda i:(f"={AR(nol_key)}" if i==0 else f"={col(i-1)}{nolbop_r+3}")); r+=1
    L(r,"NOL 사용 (80% 한도)","noluse",lambda i:f"=MIN({col(i)}{PROW['nolbop']},{col(i)}{PROW['txpre']}*{AR('nol_lim')})"); r+=1
    L(r,"NOL 추가 (당기손실)","noladd",lambda i:f"=MAX(0,-{col(i)}{PROW['ebitpre']})"); r+=1
    L(r,"NOL 기말잔액","noleop",lambda i:f"={col(i)}{PROW['nolbop']}-{col(i)}{PROW['noluse']}+{col(i)}{PROW['noladd']}"); r+=1
    L(r,"과세소득 post-NOL","txinc",lambda i:f"={col(i)}{PROW['txpre']}-{col(i)}{PROW['noluse']}"); r+=1
    L(r,"현금세금 (AMPC 비과세)","cashtax",lambda i:f"=-{col(i)}{PROW['txinc']}*{AR('tax')}"); r+=1
    L(r,"NOPAT (NOL 반영)","nopat",lambda i:f"={col(i)}{PROW['ebit']}+{col(i)}{PROW['cashtax']}",bold=True,top=True); r+=1
    L(r,"Plus: D&A","addda",lambda i:f"=-{col(i)}{PROW['da']}"); r+=1
    L(r,"Less: maint capex","mcapex",lambda i:f"=-{col(i)}{PROW['rev']}*{AR('mcapex')}"); r+=1
    L(r,"Less: growth capex","gcapex",lambda i:f"=-{AR(capex_key,i)}"); r+=1
    L(r,"Plus: CapEx 그랜트 (10%)","grant",lambda i:f"={AR(capex_key,i)}*{AR('capex_grant')}"); r+=1
    PROW['nwclvl']=r; ws.cell(r,2,"NWC level").font=fnt(10)
    for i in range(len(YEARS)):
        c=ws.cell(r,C0+i,f"={col(i)}{PROW['rev']}*{AR('nwc')}"); c.font=fnt(10); c.number_format=USD; c.alignment=right
    r+=1
    L(r,"Less: ΔNWC","dnwc",lambda i:(f"=-{col(i)}{PROW['nwclvl']}" if i==0 else f"=-({col(i)}{PROW['nwclvl']}-{col(i-1)}{PROW['nwclvl']})")); r+=1
    L(r,"Unlevered FCFF","fcff",lambda i:f"={col(i)}{PROW['nopat']}+{col(i)}{PROW['addda']}+{col(i)}{PROW['mcapex']}+{col(i)}{PROW['gcapex']}+{col(i)}{PROW['grant']}+{col(i)}{PROW['dnwc']}",bold=True,top=True,fill=fsand); r+=1
    return PROW

# ---- SKBA sheet (single pouch formfactor) ----
ba=wb.create_sheet("SKBA"); ba.sheet_view.showGridLines=False
ba.column_dimensions['A'].width=2; ba.column_dimensions['B'].width=32; ba.column_dimensions['C'].width=11
for i in range(len(YEARS)): ba.column_dimensions[col(i)].width=9.5
ba.cell(1,2,"SKBA — P&L → FCFF (4개 파우치 라인)").font=fnt(14,True,NAVY)
ba.cell(2,2,"공장1-1(L1-1A/B '27) + 공장1-2(L1-2C/D '28) · 라인 합산").font=fnt(9,False,"8C8473")
PB={}
r=4; Srow(ba,r,"REVENUE",P1C); r+=1; H(ba,r); r+=1
ba.cell(r,2,"Shipped (GWh)").font=fnt(10); PB['gwh']=r
for i in range(len(YEARS)): form(ba,r,i,f"={AR('ba_gwh',i)}",NUM1)
r+=1
ba.cell(r,2,"ASP ($/kWh)").font=fnt(10); PB['asp']=r
for i in range(len(YEARS)): form(ba,r,i,f"={AR('ba_asp',i)}",USD)
r+=1
ba.cell(r,2,"Revenue").font=fnt(10,True,NAVY); PB['rev']=r
for i in range(len(YEARS)):
    c=ba.cell(r,C0+i,f"={col(i)}{PB['gwh']}*{col(i)}{PB['asp']}"); c.font=fnt(10,True); c.number_format=USD; c.alignment=right; c.border=b_t
ba.cell(r,2).border=b_t
r+=1
ba.cell(r,2,"Cash COGS (1-1동+1-2동)").font=fnt(10); PB['cogs']=r
for i in range(len(YEARS)):
    f=f"=-({AR('ba11_gwh',i)}*{AR('ba11_cogs',i)}+{AR('ba12_gwh',i)}*{AR('ba12_cogs',i)})"
    c=ba.cell(r,C0+i,f); c.font=fnt(10); c.number_format=USD; c.alignment=right
r+=1
PB['after_rev']=r
plant_common(ba,PB,PB['rev'], lambda i:[AR('ba_gwh',i)], 'ba_capex', P1C, 'nol_ba', base_key='ba_base_ebitda')

# ---- SKOT sheet (dual formfactor: prismatic + pouch) ----
ot=wb.create_sheet("SKOT"); ot.sheet_view.showGridLines=False
ot.column_dimensions['A'].width=2; ot.column_dimensions['B'].width=32; ot.column_dimensions['C'].width=11
for i in range(len(YEARS)): ot.column_dimensions[col(i)].width=9.5
ot.cell(1,2,"SKOT — P&L → FCFF (각형 2라인 + 파우치 1라인)").font=fnt(14,True,NAVY)
ot.cell(2,2,"폼팩터별 매출 분리 · 각형 13.5×2 + 파우치 2.7 = 29.7GWh · COD '28").font=fnt(9,False,"8C8473")
PO={}
r=4; Srow(ot,r,"REVENUE — by formfactor",P2C); r+=1; H(ot,r); r+=1
ot.cell(r,2,"각형 revenue").font=fnt(10); PO['pr_rev']=r
for i in range(len(YEARS)):
    c=ot.cell(r,C0+i,f"={AR('ot_pr_gwh',i)}*{AR('ot_pr_asp',i)}"); c.font=fnt(10); c.number_format=USD; c.alignment=right; c.fill=PatternFill("solid",fgColor=OTtint)
r+=1
ot.cell(r,2,"파우치 revenue").font=fnt(10); PO['po_rev']=r
for i in range(len(YEARS)):
    c=ot.cell(r,C0+i,f"={AR('ot_po_gwh',i)}*{AR('ot_po_asp',i)}"); c.font=fnt(10); c.number_format=USD; c.alignment=right; c.fill=PatternFill("solid",fgColor=OTtint)
r+=1
ot.cell(r,2,"Revenue (total)").font=fnt(10,True,NAVY); PO['rev']=r
for i in range(len(YEARS)):
    c=ot.cell(r,C0+i,f"={col(i)}{PO['pr_rev']}+{col(i)}{PO['po_rev']}"); c.font=fnt(10,True); c.number_format=USD; c.alignment=right; c.border=b_t
ot.cell(r,2).border=b_t
r+=1
ot.cell(r,2,"Cash COGS (by formfactor)").font=fnt(10); PO['cogs']=r
for i in range(len(YEARS)):
    f=f"=-({AR('ot_pr_gwh',i)}*{AR('ot_pr_cogs',i)}+{AR('ot_po_gwh',i)}*{AR('ot_po_cogs',i)})"
    c=ot.cell(r,C0+i,f); c.font=fnt(10); c.number_format=USD; c.alignment=right
r+=1
PO['after_rev']=r
plant_common(ot,PO,PO['rev'], lambda i:[AR('ot_pr_gwh',i),AR('ot_po_gwh',i)], 'ot_capex', P2C, 'nol_ot')

# ============================================================
# CONSOLIDATED
# ============================================================
cs=wb.create_sheet("Consolidated"); cs.sheet_view.showGridLines=False
cs.column_dimensions['A'].width=2; cs.column_dimensions['B'].width=34; cs.column_dimensions['C'].width=11
for i in range(len(YEARS)): cs.column_dimensions[col(i)].width=9.5
cs.cell(1,2,"CONSOLIDATED — SKBA + SKOT → FCFF → DCF").font=fnt(15,True,NAVY)
cs.cell(2,2,"라인 단위 합산 · $mm · FCFF unlevered · AMPC 옵션A · SKOT placeholder 입력 시 합산").font=fnt(9,False,"8C8473")
CR={}
def CH(r):
    cs.cell(r,2,"($mm)").font=fnt(9,True,NAVY); cs.cell(r,2).border=b_tb; cs.cell(r,3).border=b_tb
    for i,y in enumerate(YEARS):
        c=cs.cell(r,C0+i,str(y)); c.font=fnt(10,True,NAVY); c.alignment=center; c.border=b_tb
def CL(r,label,key,fnc,bold=False,fmt=USD,top=False,fill=None,green=False,pct=False):
    lc=cs.cell(r,2,label); lc.font=fnt(10,bold,NAVY if bold else BLACK); CR[key]=r
    for i in range(len(YEARS)):
        c=cs.cell(r,C0+i,fnc(i)); c.font=fnt(10,bold,GREEN if green else BLACK); c.number_format=(PCT if pct else fmt); c.alignment=right
        if fill: c.fill=fill
        if top: c.border=b_t
    if fill: lc.fill=fill
    if top: lc.border=b_t
def CS_(r,t,fill=fnav):
    cs.cell(r,2,t).font=fnt(10,True,"FFFFFF")
    for cc in range(2,C0+len(YEARS)): cs.cell(r,cc).fill=fill
r=4
CS_(r,"REVENUE — by plant"); r+=1; CH(r); r+=1
CL(r,"SKBA revenue (파우치)","ba_rev",lambda i:f"=SKBA!{col(i)}{PB['rev']}",green=True,fill=PatternFill('solid',fgColor=BAtint)); r+=1
CL(r,"SKOT revenue (각형+파우치)","ot_rev",lambda i:f"=SKOT!{col(i)}{PO['rev']}",green=True,fill=PatternFill('solid',fgColor=OTtint)); r+=1
CL(r,"Total revenue","rev",lambda i:f"={col(i)}{CR['ba_rev']}+{col(i)}{CR['ot_rev']}",bold=True,top=True); r+=1
CL(r,"  Total shipped (GWh)","gwh",lambda i:f"=SKBA!{col(i)}{PB['gwh']}+{AR('ot_pr_gwh',i)}+{AR('ot_po_gwh',i)}",fmt=NUM1); r+=2
CS_(r,"EBIT"); r+=1; CH(r); r+=1
CL(r,"EBIT pre-AMPC","ebitpre",lambda i:f"=SKBA!{col(i)}{PB['ebitpre']}+SKOT!{col(i)}{PO['ebitpre']}"); r+=1
CL(r,"  margin pre-AMPC","ebitprem",lambda i:f"=IF({col(i)}{CR['rev']}=0,0,{col(i)}{CR['ebitpre']}/{col(i)}{CR['rev']})",pct=True); r+=1
CL(r,"Plus: AMPC 45X","ampc",lambda i:f"=SKBA!{col(i)}{PB['ampc']}+SKOT!{col(i)}{PO['ampc']}"); r+=1
CL(r,"EBIT incl. AMPC","ebit",lambda i:f"=SKBA!{col(i)}{PB['ebit']}+SKOT!{col(i)}{PO['ebit']}",bold=True,top=True); r+=1
CL(r,"  margin incl. AMPC","ebitm",lambda i:f"=IF({col(i)}{CR['rev']}=0,0,{col(i)}{CR['ebit']}/{col(i)}{CR['rev']})",pct=True); r+=1
CL(r,"Plus: D&A","da",lambda i:f"=-(SKBA!{col(i)}{PB['da']}+SKOT!{col(i)}{PO['da']})"); r+=1
CL(r,"EBITDA incl. AMPC","ebitda",lambda i:f"={col(i)}{CR['ebit']}+{col(i)}{CR['da']}",bold=True,top=True); r+=1
CL(r,"  EBITDA margin","ebitdam",lambda i:f"=IF({col(i)}{CR['rev']}=0,0,{col(i)}{CR['ebitda']}/{col(i)}{CR['rev']})",pct=True); r+=1
CL(r,"EBITDA ex-AMPC (멀티플용)","ebitda_xa",lambda i:f"={col(i)}{CR['ebitda']}-{col(i)}{CR['ampc']}",bold=True); r+=1
CL(r,"  EBITDA ex-AMPC margin","ebitda_xam",lambda i:f"=IF({col(i)}{CR['rev']}=0,0,{col(i)}{CR['ebitda_xa']}/{col(i)}{CR['rev']})",pct=True); r+=2
CS_(r,"UNLEVERED FCFF"); r+=1; CH(r); r+=1
CL(r,"SKBA FCFF","ba_fcff",lambda i:f"=SKBA!{col(i)}{PB['fcff']}",green=True); r+=1
CL(r,"SKOT FCFF","ot_fcff",lambda i:f"=SKOT!{col(i)}{PO['fcff']}",green=True); r+=1
CL(r,"Consolidated FCFF","fcff",lambda i:f"={col(i)}{CR['ba_fcff']}+{col(i)}{CR['ot_fcff']}",bold=True,top=True,fill=fsand); r+=2

# ---------- FCFE 브리지 (공장별, 링펜싱) ----------
CS_(r,"FCFE — 공장별 (하이브리드: HoldCo 우산 + ring-fence)",fred); r+=1; CH(r); r+=1
# SKBA: project debt 이자(세후) + HoldCo 크레딧비용 배분
CL(r,"SKBA · 이자비용 (project debt)","ba_int",lambda i:f"=-{AR('ba_debt')}*{AR('ba_rate')}"); r+=1
CL(r,"SKBA · 이자 tax shield","ba_shld",lambda i:f"=-{col(i)}{CR['ba_int']}*{AR('tax')}"); r+=1
CL(r,"SKBA · HoldCo 크레딧비용 배분","ba_hccost",lambda i:f"=-{AR('ba_hc')}*{AR('hc_rate')}"); r+=1
CL(r,"SKBA FCFE","ba_fcfe",lambda i:f"={col(i)}{CR['ba_fcff']}+{col(i)}{CR['ba_int']}+{col(i)}{CR['ba_shld']}+{col(i)}{CR['ba_hccost']}",bold=True,top=True); r+=1
# SKOT: DOE 이자(세후) + 메자닌(amortizing) + HoldCo 크레딧비용 배분
CL(r,"SKOT · DOE 이자 (3.5%)","ot_int",lambda i:f"=-{AR('doe_bal')}*{AR('doe_rate')}"); r+=1
CL(r,"SKOT · DOE 이자 tax shield","ot_shld",lambda i:f"=-{col(i)}{CR['ot_int']}*{AR('tax')}"); r+=1
# 메자닌 amortizing: 상환개시~만기 균등상환 (현금발생後 거치), 잔액 직접계산
CL(r,"메자닌 연상환액","mez_pmt",lambda i:f"={AR('pref_bal')}/({AR('pref_mat')}-{AR('pref_amort_start')}+1)"); r+=1
CL(r,"메자닌 기초잔액","mez_bop",lambda i:f"=MAX(0,{AR('pref_bal')}-{col(i)}{CR['mez_pmt']}*MAX(0,MIN(({YEARS[i]})-{AR('pref_amort_start')},{AR('pref_mat')}-{AR('pref_amort_start')}+1)))"); r+=1
CL(r,"메자닌 원금상환 (amort, AMPC재원)","mez_amort",lambda i:f"=-IF(AND(({YEARS[i]})>={AR('pref_amort_start')},({YEARS[i]})<={AR('pref_mat')}),MIN({col(i)}{CR['mez_bop']},{col(i)}{CR['mez_pmt']}),0)"); r+=1
CL(r,"SKOT · 메자닌 배당 (7%, 잔액기준)","ot_pref",lambda i:f"=-{col(i)}{CR['mez_bop']}*{AR('pref_rate')}"); r+=1
CL(r,"SKOT · HoldCo 크레딧비용 배분","ot_hccost",lambda i:f"=-{AR('ot_hc')}*{AR('hc_rate')}"); r+=1
CL(r,"SKOT FCFE (보통주 귀속)","ot_fcfe",lambda i:f"={col(i)}{CR['ot_fcff']}+{col(i)}{CR['ot_int']}+{col(i)}{CR['ot_shld']}+{col(i)}{CR['ot_pref']}+{col(i)}{CR['mez_amort']}+{col(i)}{CR['ot_hccost']}",bold=True,top=True); r+=1
# DSCR (메자닌+DOE 관점)
CL(r,"OT CFADS (EBITDA-유지capex)","ot_cfads",lambda i:f"=SKOT!{col(i)}{PO['ebit']}-SKOT!{col(i)}{PO['da']}+SKOT!{col(i)}{PO['mcapex']}",green=True); r+=1
CL(r,"  부채상환액 (DOE이자+메자닌 원리금)","ot_ds",lambda i:f"=-{col(i)}{CR['ot_int']}-{col(i)}{CR['ot_pref']}-{col(i)}{CR['mez_amort']}"); r+=1
CL(r,"  DSCR (CFADS÷상환액)","ot_dscr",lambda i:f"=IF({col(i)}{CR['ot_ds']}=0,0,{col(i)}{CR['ot_cfads']}/{col(i)}{CR['ot_ds']})",fmt='0.00"x"'); r+=1
# 합산 보통주(SK) FCFE
CL(r,"보통주(SK) FCFE 합계","fcfe",lambda i:f"={col(i)}{CR['ba_fcfe']}+{col(i)}{CR['ot_fcfe']}",bold=True,top=True,fill=fsand); r+=1
CL(r,"FCFE 할인계수 (Ke)","dfe",lambda i:f"=1/(1+{AR('coe')})^(({i+1})-{AR('midyr')})",fmt='0.000'); r+=1
CL(r,"PV of FCFE","pve",lambda i:f"={col(i)}{CR['fcfe']}*{col(i)}{CR['dfe']}",bold=True); r+=1
cs.cell(r,2,"※ 메자닌 제거: OT 소요 $1bn은 BA 증자(equity-down)로 조달 → OT 신규부채·DSCR 불요. 수익률은 보통주 FCFE(Ke)로 평가. DOE 이자만·각 SPV non-recourse.").font=fnt(8,False,"8C8473",it=True); r+=2

CS_(r,"DCF VALUATION",fred); r+=1; CH(r); r+=1
CL(r,"Discount period (mid-yr)","disc",lambda i:f"=({i+1})-{AR('midyr')}",fmt='0.0'); r+=1
CL(r,"Discount factor","df",lambda i:f"=1/(1+{AR('wacc')})^{col(i)}{CR['disc']}",fmt='0.000'); r+=1
CL(r,"PV of FCFF","pv",lambda i:f"={col(i)}{CR['fcff']}*{col(i)}{CR['df']}",bold=True); r+=2
last=col(len(YEARS)-1)
cs.cell(r,2,"VALUATION SUMMARY ($mm)").font=fnt(11,True,"FFFFFF")
for cc in range(2,6): cs.cell(r,cc).fill=fred
r+=1
def vr(r,label,f,bold=False,fill=None):
    cs.cell(r,2,label).font=fnt(11 if bold else 10,bold,NAVY if bold else BLACK)
    c=cs.cell(r,4,f); c.font=fnt(11 if bold else 10,bold,BLACK); c.number_format=USD; c.alignment=right
    if fill: cs.cell(r,2).fill=fill; c.fill=fill; cs.cell(r,2).border=b_tb; c.border=b_tb
    return r
PVEXP=vr(r,"Sum PV explicit FCFF (2026–35)",f"=SUM({col(0)}{CR['pv']}:{last}{CR['pv']})"); r+=1
TV=vr(r,"Terminal value (Gordon, undisc.)",f"={last}{CR['fcff']}*(1+{AR('tgr')})/({AR('wacc')}-{AR('tgr')})"); r+=1
PVTV=vr(r,"PV of terminal value",f"=D{TV}*{last}{CR['df']}"); r+=1
EV=vr(r,"Enterprise Value (EV)",f"=D{PVEXP}+D{PVTV}",bold=True,fill=fsand); r+=1
vr(r,"Less: net debt (DOE loan)",f"=-{AR('netdebt')}"); r+=1
cs.cell(r,2,"Equity Value (DCF)").font=fnt(11,True,"FFFFFF"); cs.cell(r,2).fill=fred
c=cs.cell(r,4,f"=D{EV}-{AR('netdebt')}"); c.font=fnt(11,True,"FFFFFF"); c.number_format=USD; c.alignment=right; c.fill=fred
EQ=r
r+=2
cs.cell(r,2,"% TV of EV").font=fnt(9,False,"8C8473")
cc=cs.cell(r,4,f"=D{PVTV}/D{EV}"); cc.font=fnt(9,False,"8C8473"); cc.number_format=PCT; cc.alignment=right
r+=2
# FCFE-based Equity Value (직접법)
cs.cell(r,2,"EQUITY VALUE — FCFE 직접법 ($mm)").font=fnt(11,True,"FFFFFF")
for cc in range(2,6): cs.cell(r,cc).fill=fred
r+=1
PVEXPE=vr(r,"Sum PV explicit FCFE (2026–35)",f"=SUM({col(0)}{CR['pve']}:{last}{CR['pve']})"); r+=1
TVE=vr(r,"FCFE Terminal value (Gordon)",f"={last}{CR['fcfe']}*(1+{AR('tgr')})/({AR('coe')}-{AR('tgr')})"); r+=1
PVTVE=vr(r,"PV of FCFE terminal value",f"=D{TVE}*1/(1+{AR('coe')})^(({len(YEARS)})-{AR('midyr')})"); r+=1
EQE=vr(r,"Equity Value (FCFE 직접법)",f"=D{PVEXPE}+D{PVTVE}",bold=True,fill=fsand); r+=1
cs.cell(r,2,"vs. Equity Value (DCF: EV−순차입금)").font=fnt(9,False,"8C8473")
c=cs.cell(r,4,f"=D{EQ}"); c.font=fnt(9,False,"8C8473"); c.number_format=USD; c.alignment=right; r+=1
cs.cell(r,2,"※ FCFE는 Ke로 할인(레버리지 직접 반영) / DCF는 FCFF를 WACC로 할인 후 순차입금 차감. 두 경로의 주주가치 교차검증.").font=fnt(8,False,RED,it=True)
r+=2
# SOTP cross-check (NAV) — NAV는 순자산(=자본총계)이라 부채 이미 차감됨 → Equity Value와 비교
cs.cell(r,2,"SOTP 교차검증 — 순자산 NAV vs 지분가치 ($mm)").font=fnt(10,True,"FFFFFF")
for cc2 in range(2,6): cs.cell(r,cc2).fill=fnav
r+=1
cs.cell(r,2,"SKBA 순자산 (DART, 부채차감後)").font=fnt(10)
c=cs.cell(r,4,f"={AR('skba_nav')}"); c.font=fnt(10,False,GREEN); c.number_format=USD; c.alignment=right; r+=1
cs.cell(r,2,"SKOT 순자산 (분할 PPA, DOE차감後)").font=fnt(10)
c=cs.cell(r,4,f"={AR('skot_nav')}"); c.font=fnt(10,False,GREEN); c.number_format=USD; c.alignment=right; r+=1
cs.cell(r,2,"합산 NAV = SOTP 지분가치").font=fnt(10,True,NAVY); cs.cell(r,2).fill=fsand
c=cs.cell(r,4,f"={AR('skba_nav')}+{AR('skot_nav')}"); c.font=fnt(10,True); c.number_format=USD; c.alignment=right; c.fill=fsand; c.border=b_tb; cs.cell(r,2).border=b_tb; r+=1
cs.cell(r,2,"vs. DCF Equity Value").font=fnt(10,True,NAVY)
c=cs.cell(r,4,f"=D{EQ}"); c.font=fnt(10,True,RED); c.number_format=USD; c.alignment=right; r+=1
cs.cell(r,2,"  괴리 (NAV − DCF Equity)").font=fnt(9,False,"8C8473")
c=cs.cell(r,4,f"=({AR('skba_nav')}+{AR('skot_nav')})-D{EQ}"); c.font=fnt(9,False,"8C8473"); c.number_format=USD; c.alignment=right; r+=1
cs.cell(r,2,"※ NAV는 순자산(자본총계)이라 부채 이미 차감 → 순차입금 재차감 不可. DCF는 EV에서만 순차입금 1회 차감. 둘 다 '지분가치' 기준 동일 비교.").font=fnt(8,False,RED,it=True)
r+=1
cs.cell(r,2,"괴리 의미: NAV(장부 순자산) ≫ DCF Equity → 사업 현금흐름이 장부 순자산·DOE 부채를 정당화 못함 (자산>수익력).").font=fnt(8,False,"8C8473",it=True)
r+=2
# EV/EBITDA forward multiple valuation
cs.cell(r,2,"EV/EBITDA 멀티플 밸류에이션 (Forward, 증권사 추정 기준)").font=fnt(10,True,"FFFFFF")
for cc2 in range(2,6): cs.cell(r,cc2).fill=fred
r+=1
ebitda_2029=f"{col(3)}{CR['ebitda_xa']}"  # 2029 ex-AMPC EBITDA
ebitda_rr=f"AVERAGE({col(4)}{CR['ebitda_xa']}:{col(6)}{CR['ebitda_xa']})"  # 2030-32 run-rate ex-AMPC
cs.cell(r,2,"Forward EBITDA ex-AMPC (2029E)").font=fnt(10)
c=cs.cell(r,4,f"={ebitda_2029}"); c.font=fnt(10,False,GREEN); c.number_format=USD; c.alignment=right; FWE=r; r+=1
cs.cell(r,2,"Run-rate EBITDA ex-AMPC (2030–32 평균)").font=fnt(10)
c=cs.cell(r,4,f"={ebitda_rr}"); c.font=fnt(10,False,GREEN); c.number_format=USD; c.alignment=right; RRE=r; r+=1
cs.cell(r,2,"peer forward EV/EBITDA 가정 (×)").font=fnt(9,True,NAVY); cs.cell(r,2).border=b_t
for j,m in enumerate([8,11,14]):
    cc=cs.cell(r,4+j,m); cc.font=fnt(10,True,BLUE); cc.number_format='0"x"'; cc.alignment=center; cc.border=b_t
MULT=r; r+=1
cs.cell(r,7,"보수 / 기준 / 강세").font=fnt(8,False,"8C8473",it=True); r+=1
cs.cell(r,2,"→ 내재 EV (2029E ex-AMPC EBITDA × 배수)").font=fnt(10,True,NAVY)
for j in range(3):
    cc=cs.cell(r,4+j,f"=$D${FWE}*{get_column_letter(C0+j)}{MULT}")
    cc.font=fnt(10,True,RED); cc.number_format=USD; cc.alignment=right
r+=1
cs.cell(r,2,"→ 내재 EV (run-rate ex-AMPC EBITDA × 배수)").font=fnt(10,True,NAVY)
for j in range(3):
    cc=cs.cell(r,4+j,f"=$D${RRE}*{get_column_letter(C0+j)}{MULT}")
    cc.font=fnt(10,True,RED); cc.number_format=USD; cc.alignment=right
r+=1
cs.cell(r,2,"vs. DCF EV (관세하한)").font=fnt(9,False,"8C8473")
c=cs.cell(r,4,f"=D{EV}"); c.font=fnt(9,False,"8C8473"); c.number_format=USD; c.alignment=right; r+=1
cs.cell(r,2,"※ EBITDA에서 AMPC 제외(2033 소멸 한시 보조금이라 영구 자본화 부적절) · EBIT 마진 10% 목표 · peer forward 8~14x.").font=fnt(8,False,RED,it=True)
r+=1
cs.cell(r,2,"  단, peer(LGES·삼성SDI·Fluence) EBITDA에도 AMPC 포함 → 우리만 ex-AMPC면 보수적 비교(apples-to-apples 아님 유의).").font=fnt(8,False,"8C8473",it=True)
r+=2
# chart
from openpyxl.chart import BarChart, LineChart, Reference
cat=Reference(cs,min_col=C0,max_col=C0+len(YEARS)-1,min_row=5,max_row=5)
bar=BarChart(); bar.type="col"; bar.title="Revenue by plant & Consolidated FCFF ($mm)"; bar.height=8; bar.width=22; bar.grouping="stacked"; bar.overlap=100
for k in ('ba_rev','ot_rev'):
    bar.add_data(Reference(cs,min_col=C0-2,max_col=C0+len(YEARS)-1,min_row=CR[k],max_row=CR[k]),titles_from_data=True,from_rows=True)
bar.set_categories(cat); bar.x_axis.delete=False; bar.y_axis.delete=False
ln=LineChart(); ln.add_data(Reference(cs,min_col=C0-2,max_col=C0+len(YEARS)-1,min_row=CR['fcff'],max_row=CR['fcff']),titles_from_data=True,from_rows=True); ln.set_categories(cat)
bar+=ln
cs.add_chart(bar,f"B{r+3}")

import os
_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
os.makedirs(_OUT, exist_ok=True)
wb.save(os.path.join(_OUT, "FCFF_model_lines.xlsx"))
print("saved line-level model")
