"""역산 밸류에이션 시나리오 — 조달 비중(그랜트/CoC 희석)이 implied EV/EBITDA 멀티플에 미치는 영향.
run-rate ex-AMPC EBITDA는 재계산된 모델(FCFF_model_lines.xlsx)에서 읽어온다."""
import os, sys
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "output", "FCFF_model_lines.xlsx")

# --- 모델에서 run-rate ex-AMPC EBITDA 읽기 ---
wb = load_workbook(XLSX, data_only=True)
cs = wb["Consolidated"]
def findrow(text):
    for row in cs.iter_rows(min_col=2, max_col=2):
        if row[0].value and text in str(row[0].value):
            return row[0].row
RRE = cs.cell(findrow("Run-rate EBITDA ex-AMPC"), 4).value  # $mm
EV_dcf = cs.cell(findrow("Enterprise Value (EV)"), 4).value

# --- 고정 파라미터 ---
RAISE   = 1420.0   # 신규 growth capex 조달총액 ($mm) = BA 일괄조달
NETDEBT = 6400.0   # BA/OT 합산 순차입금 (고정)
DEBT_PCT = 0.60    # 비소구부채 비중(고정) — 그랜트↑ 시 지분(equity)이 그만큼 감소
PEER_LO, PEER_HI = 8.0, 13.0

def scen(name, grant_pct, dilution_cap):
    equity_pct = 1.0 - grant_pct - DEBT_PCT     # 지분 = 100% - 그랜트 - 부채
    ext_equity = equity_pct * RAISE             # 외부 지분 체크
    min_equity = ext_equity / dilution_cap      # 희석 한도 지키는 최소 지분가치
    min_ev     = min_equity + NETDEBT
    mult       = min_ev / RRE
    inpeer     = "IN peer" if mult <= PEER_HI else f"+{mult-PEER_HI:.1f}x over"
    print(f"{name:<34} grant {grant_pct:>3.0%} | equity {equity_pct:>3.0%} | dil {dilution_cap:>3.0%} "
          f"| ext-eq ${ext_equity:>4.0f} | min-eq ${min_equity:>5.0f} | min-EV ${min_ev:>5.0f} "
          f"| {mult:>5.1f}x  [{inpeer}]")
    return mult

print(f"run-rate ex-AMPC EBITDA (모델)   = ${RRE:.0f}mm")
print(f"DCF EV (모델)                    = ${EV_dcf:.0f}mm")
print(f"고정: 조달총액 ${RAISE:.0f}mm · 순차입금 ${NETDEBT:.0f}mm · 부채비중 {DEBT_PCT:.0%} · peer {PEER_LO:.0f}~{PEER_HI:.0f}x\n")

print("=== 명명 시나리오 ===")
scen("BASE (grant10/eq30/dil20)",        0.10, 0.20)
scen("① CoC 20% 유지 + 그랜트 20%",       0.20, 0.20)
scen("② CoC 해제 → 30% 희석 (grant10)",   0.10, 0.30)
scen("①+② 결합 (grant20 + dil30)",        0.20, 0.30)

print("\n=== 민감도 매트릭스 : implied min EV/EBITDA (x) ===")
grants = [0.10, 0.15, 0.20]
dils   = [0.20, 0.25, 0.30]
hdr = "grant\\dil " + "".join(f"{d:>8.0%}" for d in dils)
print(hdr)
for g in grants:
    row = f"  {g:>5.0%}   "
    for d in dils:
        eq = (1.0 - g - DEBT_PCT) * RAISE
        m  = (eq/d + NETDEBT) / RRE
        row += f"{m:>7.1f}x"
    print(row)
print(f"\n(peer forward EV/EBITDA {PEER_LO:.0f}~{PEER_HI:.0f}x — 매트릭스 값이 이 밴드 안이면 딜 성립 여력)")
